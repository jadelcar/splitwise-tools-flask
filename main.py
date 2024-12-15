import uvicorn
import random

from fastapi import FastAPI, HTTPException, Depends, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.requests import Request
from fastapi.testclient import TestClient
from fastapi.templating import Jinja2Templates

from sqlalchemy.orm import Session

import starlette.status as status
from starlette.middleware import Middleware
from starlette.middleware.sessions import SessionMiddleware

from splitwise import Splitwise
from splitwise.expense import Expense
from splitwise.user import ExpenseUser

from decimal import *
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side

# Internal imports

from core.constants import *
from core.config.settings import get_settings
from api.routes import auth
from core.exceptions import handlers
from services.helpers.expense_utils import *

# from db.database import engine, SessionLocal, database.get_db
from db import crud, models, schemas, database

# Set up FastAPI
middleware = [
    Middleware(SessionMiddleware, secret_key='super-secret')
]
app = FastAPI(middleware=middleware)

# Configure templating
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="static/templates")

settings = get_settings()
URL = f"http://{settings.APP_HOST}:{settings.APP_PORT}"
CONSUMER_KEY = settings.CONSUMER_KEY
CONSUMER_SECRET = settings.CONSUMER_SECRET


# Configure database
models.Base.metadata.create_all(bind = database.engine) # Creates DB if not yet created


"""       ----------           PATHS       -----------            """

@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    """Return home page """
    return templates.TemplateResponse("home.html", {"request": request})


"""       ----------           Authentication       -----------            """

@app.get("/login_sw")
def login_sw(request: Request):
    sObj = Splitwise(CONSUMER_KEY,  CONSUMER_SECRET)
    url, state = sObj.getOAuth2AuthorizeURL(URL + "/authorize") 
    
    request.session['state'] = state # Store state in session to double check later
    
    return RedirectResponse(url) #Redirect user to SW authorization website. After login, redirects user to the URL defined in the app's settings

@app.get("/authorize", response_class = HTMLResponse)
def authorize(request: Request, code: str, state: str):
    """
    The user is redirected here after granting access to the app in Splitwise.
    
    Parameters:
    code (str): authorization code received from SW
    state (str): state that was sent in the initial request to SW
    
    Returns:
    HTMLResponse: redirects user to 'authorize_success.html'
    """
    
    # Get parameters needed to obtain the access token
    sObj = Splitwise(CONSUMER_KEY, CONSUMER_SECRET)
    
    # Check that state is the same
    state_previous = request.session.get('state')
    if state_previous != state:
        raise Exception("State is not the same")

    access_token = sObj.getOAuth2AccessToken(code, URL + "/authorize")
    sObj.setOAuth2AccessToken(access_token)
    
    # Store user data and tokens in session
    request.session['access_token'] = access_token
    current_user = sObj.getCurrentUser()
    request.session['user_id'] = current_user.id
    request.session['user_fname'] = current_user.first_name
    return templates.TemplateResponse("authorize_success.html", {"request": request})

@app.get("/logout")
def logout(request: Request):
    """
    Logout user by clearing the session and redirecting to the home page.
    
    Returns:
    RedirectResponse: redirects user to the home page
    """
    request.session.clear()
    return RedirectResponse("/")


"""       ----------           Retrieve data       -----------            """

@app.get('/batch_upload', response_class = HTMLResponse)
def batch_upload_show_form(request: Request):
    """
    Show form for uploading an excel file
    """
    sObj = auth.get_access_token(request)
    groups = sObj.getGroups()
    return templates.TemplateResponse("batch_upload.html", {"request": request, "groups" : groups})


@app.post("/batch_upload_process", response_class = HTMLResponse)
def batch_upload_process(request: Request, db: Session = Depends(database.get_db),group_for_upload = Form(), batch_expenses_file : UploadFile = File(...)):
    """
    Process the data uploaded
    """
    # Access tokens
    sObj = auth.get_access_token(request)
    current_user = sObj.getCurrentUser()
    
    group = sObj.getGroup(group_for_upload) # Fetch group info
    request.session.pop('upload_id', None) # Erase from session any previous upload
    file = batch_expenses_file.file.read() # Import and parse user file 
    
    expenses_df = pd.read_excel(file, sheet_name = "Expenses") # Import expenses sheet
    members_df = pd.read_excel(file, sheet_name = "Members") # Import the members sheet
    
    # Parse the expenses sheet
    cols_member_names = list(expenses_df.filter(regex='^_', axis=1)) # Make a list with member names
    expenses_df['Total Shares'] = expenses_df[cols_member_names].sum(axis = 1) # Create col to add shares from all members
    
    expenses_df['All equal'] = expenses_df['All equal'].astype(str).apply(str.lower).replace(['y','n',''], [True, False, False]) # Read the 'All equal' columns
        
    # Create a list of dicts {name: , id:} to hold member info (name and ID)
    members_in_cols = []
    for member_name in expenses_df[cols_member_names].columns:
        member_info = members_df[members_df['Name'] == member_name[1:]]
        members_in_cols.append(
            {
                'name': member_name[1:],
                'id': int(member_info['ID'].values[0])
            }
        )

    # Payer ID
    expenses_df = expenses_df.merge(members_df, 
                                    how = 'left', 
                                    left_on = "Paid by", 
                                    right_on = 'Name').rename(
                                        columns = {'ID_x': "ID", 'ID_y': "Payer ID"}
                                        )

    # Calculate share owed and paid for a given member
    def getShareOwed(row, member_name, cols_member_names):
        """ Calculate share owed for a given expense (row) and user (member name), taking into account split type"""
        member_cell_value = row[member_name]
        if pd.isna(member_cell_value):
            return 0
        elif row['All equal']:
            return round(row['Amount'] / len(group.members), 2) # Divide equally by the number of members in group
        elif row["Split type"] == "share":
            return round(member_cell_value / 100 * row['Amount'], 2) # Divide based on %
        elif row["Split type"] == "amount":
            return member_cell_value # Assign the amount specified
        elif row["Split type"] == "equal":
            members_for_division = [member['name'] for member in members_in_cols if not pd.isna(row[f"_{member['name']}"])]
            return round(row['Amount'] / len(members_for_division), 2)
        else:
            return 0 # An error will be raise to the user
        
    for col_name in cols_member_names:
        # Share paid
        expenses_df[f'{col_name[1:]}_share_paid'] = np.where(expenses_df["Paid by"] == col_name[1:],  expenses_df['Amount'], 0)

        # Share owed
        expenses_df[f'{col_name[1:]}_share_owed'] = expenses_df.apply(getShareOwed, axis = 1, member_name = f"_{col_name[1:]}", cols_member_names=cols_member_names)

    # Round share owed if it doesn't add up
    def AssignRoundingDiff(row):
        """Assign the rounding difference to a random member within the expense
        Add up share_owed of all members and compare with total amount: If the difference is due to rounding (<0.02), subtract this from a random user within the expense
        """
        share_owed_columns = [f"{col_name[1:]}_share_owed" for col_name in cols_member_names if row[f"{col_name[1:]}_share_owed"] > 0] # List of columns to parse in this expense
        sum_share_owed = row[share_owed_columns].sum()
        diff = sum_share_owed - row['Amount']
        if abs(diff) > 0 and abs(diff) < 0.02:
            diff_round = round(diff, 2)
            random_member = random.choice(share_owed_columns)
            row[random_member] += -diff_round # Subtract the difference
        return row

    expenses_df = expenses_df.apply(AssignRoundingDiff, axis = 1)

    # Obtain error message
    errors, error_messages, error_count = describe_errors(expenses_df, members_df, group)

    # Store data temporarily so it can be pushed later
    expenses = expenses_df.to_dict('records')
    if error_count == 0:

        # Create upload and expenses in database
        new_upload = crud.create_upload(db, creator_user_id = current_user.id, group_id = group.id)
             
        for exp in expenses:
            crud.create_expense(db, upload_id = new_upload.id, expense = exp, group_members = members_in_cols, creator_user_id = current_user.id)
    
    # Prepare context to be passed to template
    context = {
        "request" : request,
        "group" : group_to_dict(group), 
        "members_in_cols" : members_in_cols, 
        "expenses" : expenses,
        "errors" : errors,
        "error_messages" : error_messages,
        "file_valid" : error_count == 0,
        "upload_id" : new_upload.id if error_count==0 else 0,
    }

    return templates.TemplateResponse("upload_edit.html", context)

@app.post('/push_expenses/{upload_id}', response_class = HTMLResponse)
def push_expenses(request: Request, upload_id: int, db: Session = Depends(database.get_db)):

    sObj = auth.get_access_token(request)
        
    upload = crud.get_upload_by_id(db, upload_id = upload_id)
    expenses = crud.get_expenses_by_upload_id(db, upload_id = upload_id)
    group = crud.get_group_by_id(db, group_db_id = upload.group_id) # DB ID, not SW ID

    # Upload each expense to Splitwise
    expenses_to_push = {}
    for e in expenses:
        expense = Expense()
        expense.setCost(e.amount)
        expense.setDescription(e.description)
        expense.setGroupId(group.sw_id)
        expense.setCreationMethod('Splitwise tools')
        expense.setDate(e.date)
        expense.setCurrencyCode(e.currency)
        # If we split all equal, no need to add members
        if e.all_equal == True and e.payer_id == sObj.getCurrentUser().getId(): 
            expense.setSplitEqually()
        # O/w, add each member of the expense
        else:
            for member in e.expense_members:
                user = ExpenseUser()
                user.setId(member.member.sw_id)
                user.setPaidShare(member.share_paid) 
                user.setOwedShare(member.share_owed)
                expense.addUser(user)

        expenses_to_push[e.within_upload_id] = expense # Add to dictionary, using 'Id' (specified by user) as key
    
    # Upload expenses
    expenses_failed = []
    for expense_id, expense in expenses_to_push.items():
        nExpense, errors = sObj.createExpense(expense)
        if nExpense is not None: 
            print("Expense ID in Splitwise: " + str(nExpense.getId()))
        else:
            errors_list = errors.getErrors()['base']
            print(f"Expense errors: {errors_list}")
            expenses_failed.append(
                {
                    'id': str(expense_id),
                    'errors' : errors_list
                }
            )

    context = {
        "request": request,
        "group" :   {   
                        'id': group.sw_id,
                        'name': group.name,
                    },
        "expenses_failed" : expenses_failed,
    }

    # Return summary
    return templates.TemplateResponse("upload_summary.html", context)

"""       ----------           Create data       -----------            """
@app.get("/create_upload", response_class=HTMLResponse)
def create_upload(request: Request, db: Session = Depends(database.get_db)):
    """Create a new upload"""
    try:
        sObj = auth.get_access_token(request)
        current_user = sObj.getCurrentUser().getId()
    except:
        current_user = 7357
    new_upload = crud.create_upload(db, creator_user_id = current_user)
    return templates.TemplateResponse("home.html", {"request": request, "new_upload" : new_upload})

@app.post("/create_expense", response_class=HTMLResponse)
def create_expense(request: Request, db: Session = Depends(database.get_db)):
    """Create a new expense"""
    try:
        sObj = auth.get_access_token(request)
        current_user = sObj.getCurrentUser().getId()
    except:
        current_user = 7357
    crud.create_expense(db, creator_user_id = current_user)
    return templates.TemplateResponse("home.html", {"request": request})


"""       ----------           Retrieve data       -----------            """

@app.get("/uploads", response_class=HTMLResponse)
def get_groups_by_id(request: Request, db: Session = Depends(database.get_db)):
    """Get uploads of the current user logged in the app, using it's user ID"""
    sObj = auth.get_access_token(request)
    sObj.getCurrentUser().id
    uploads = crud.get_uploads(db)
    return templates.TemplateResponse("uploads.html", {"request": request, "uploads" : uploads})

@app.get("/groups", response_class=HTMLResponse)
def get_groups_by_id(request: Request):
    """Get groups of the current user logged in the app, using it's app user ID"""
    sObj = auth.get_access_token(request)
    groups = sObj.getGroups()
    return templates.TemplateResponse("groups.html", {"request": request, "groups" : groups})

@app.get("/template/{group_id}")
def get_template_by_group_id(request: Request, group_id: int):
    """Create a template excel for a group"""
    sObj = auth.get_access_token(request)
    group = sObj.getGroup(group_id)
    members = group.members
    
    # Define headers    
    expense_headers = ["ID","Description","Date","Amount","Currency","Paid by","All equal","Split type"]
    members_headers = ["Name", "ID"]
    
    # List of member names
    member_names = []
    for member in members:
        member_last_name = "" if member.last_name == None else member.last_name
        member_full_name = f"{member.first_name} {member_last_name}".strip()
        member_names.append(member_full_name)

    # For repeated names, enumerate them 1-N (e.g. Javier 1, Javier 2...)
    member_names_enum = []
    for i, name in enumerate(member_names): 
        count = member_names.count(name)
        if count > 1:
            member_names_enum.append(name + " " + str(member_names[0:i + 1].count(name)))
        else:
            member_names_enum.append(name)

    # Append to headers
    expense_headers += [f"_{name}" for name in member_names_enum] # add prefix "_"
    
    # Create and configure excel file
    wb = Workbook()
    expenses_ws = wb.active
    expenses_ws.title = "Expenses"
    expenses_ws.sheet_properties.tabColor = "00cc00"

    members_ws = wb.create_sheet(title="Members")
    members_ws.sheet_properties.tabColor = "009933"

    # Create headers
    for sheet_headers, sheet in zip([expense_headers, members_headers], [expenses_ws, members_ws]):
        for c, header in enumerate(sheet_headers, start=1):
            cell = sheet.cell(row=1, column=c)
            cell.value = header
            cell.font = Font(bold=True)
            cell.border = Border(bottom=Side(style='thick', color="000000"))

    # Add member data in 'members' sheet
    for r, (member, member_name) in enumerate(zip(members, member_names_enum), start = 2):
        members_ws.cell(row = r, column = 1, value = member_name)
        members_ws.cell(row = r, column = 2, value = member.id)

    # Save and return file
    file_path = f"static/assets/group_templates/{group_id}.xlsx"
    wb.save(file_path)

    headers = {'Content-Disposition': f'attachment; filename="template-{group.name}.xlsx"'}
    return FileResponse(file_path, headers=headers, media_type = "application/vnd.ms-excel")


"""       ----------           User registration       -----------            """

@app.get("/register/", response_class = HTMLResponse)
def register_show_form(request: Request):
    """ Show user a page for registering"""
    return templates.TemplateResponse("register.html", {"request": request})

@app.post("/register_submit/")
def register(request: Request, username: str = Form(), password: str = Form(), confirmation: str = Form(), db: Session = Depends(database.get_db)):
    """ Register user using data in their registration form"""
    
    # Obtain user's data entered in the form
    username = request.get("username")
    password = request.get("password")
    confirmation = request.get("confirmation") 
    
    # Check if the username is being used
    db_user_byusername = crud.get_user_by_username(db, username=username)
    if db_user_byusername:
        return handlers.apology("That username is taken...", request, 400)

    # Validate data
    elif username=="" or password=="" or confirmation=="":
        # If at least one of the above are missing
        return handlers.apology("You didn't complete all the fields, right?", request, 400)
    elif password != confirmation :
        return handlers.apology("Passwords don't match bro", request, 400)

    # Create user in database
    try:
        new_user = schemas.UserCreate(username = username, password = password)
        crud.create_user(db, new_user)
        new_db_user = crud.get_user_by_username(db, username)
        request.session['user_id'] = new_db_user.id # Store the user id in session 
        request.session['username'] = new_db_user.username # Store the user id in session 
    except:
        return handlers.apology("Could not insert you in the database", request, 400)

    # Redirect to home
    return RedirectResponse('/', status_code=status.HTTP_302_FOUND) # See https://stackoverflow.com/a/65512571/19667698

@app.get("user/{username}")
def get_user_byusername(username: int, db: Session = Depends(database.get_db)):
    """Get user by its username"""
    user = crud.get_user_by_username(db, username = username).first()
    return user

""" Create users and groups in database"""
@app.post("/users/", response_model=schemas.User)
def create_user(user: schemas.UserCreate, db: Session = Depends(database.get_db)):
    """Create a new user in database"""
    db_user = crud.get_user_by_email(db, email=user.email)
    if db_user: #If db_user exists (i.e. the search by email return something)
        raise HTTPException(status_code=400, detail="Email already registered")
    return crud.create_user(db=db, user=user)





# Add test client and first test
client = TestClient(app)


def test_read_main():
    response = client.get("/")
    assert response.status_code == 200
    assert response.json() == {"msg": "Hello World"}


if __name__ == "__main__":
    settings = get_settings()
    uvicorn.run(
        "main:app",
        host=settings.APP_HOST,
        port=settings.APP_PORT,
        reload=True,
        reload_dirs=["db",""]
    )
